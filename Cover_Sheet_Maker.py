import os
from openpyxl import load_workbook

# Create a new folder for cover sheets
cover_sheets_folder = 'Publisher_Statements_Cover_Sheets'
os.makedirs(cover_sheets_folder, exist_ok=True)

# Process each summary sheet
summaries_folder = 'Publisher_Statement_Summaries'
for filename in os.listdir(summaries_folder):
    summary_path = os.path.join(summaries_folder, filename)

    # Open the summary sheet
    summary = load_workbook(summary_path, data_only=True)
    summary_sheet = summary.active

    # Create a new cover sheet using the template
    cover_sheet_path = os.path.join(cover_sheets_folder, f'{filename[:-5]}-CoverSheet.xlsx')
    template_path = 'CoverSheet_Template.xlsx'
    template = load_workbook(template_path)
    cover_sheet = template.active

    # Populate cover sheet with data from summary sheet
    for row in range(2, summary_sheet.max_row + 1):
        publisher_cell = summary_sheet.cell(row=row, column=1)
        balance_cell = summary_sheet.cell(row=row, column=15)  # Assuming balance column is column O (index 15)

        # Cells A23, A24, A25, ...
        cover_sheet.cell(row=row + 21, column=1, value=publisher_cell.value)
        # Cells F23, F24, F25, ...
        cover_sheet.cell(row=row + 21, column=6, value=balance_cell.value)

    # Save the cover sheet
    template.save(cover_sheet_path)

    # Close the summary workbook
    summary.close()